import torch
import torch.nn as nn
import math
import openpyxl
import torch.nn.functional as F
import os

class HashMapping(nn.Module):

    def __init__(self, input_dim: int, hidden_dim: int, output_dim: int, dropout: float):
        super().__init__()
        self.hash_layers = nn.Sequential(
            nn.Linear(input_dim, hidden_dim),
            # nn.GELU(),
            nn.ReLU(),
            nn.Dropout(p=dropout),
            nn.Linear(hidden_dim, hidden_dim),
            # nn.GELU(),
            nn.ReLU(),
            nn.Dropout(p=dropout),
            nn.Linear(hidden_dim, output_dim),
        )
        self.scale = nn.Parameter(torch.Tensor([1.0]))
        self.last_layer = nn.Tanh()
        self.weights = self.init_weights()

    def init_weights(self):
        for m in self.modules():
            if isinstance(m, nn.Linear):
                nn.init.xavier_normal_(m.weight)
                nn.init.constant_(m.bias, 0.0)

    def forward(self, text_features):
        t = self.hash_layers(text_features)
        t = self.last_layer(self.scale * t)
        b = torch.sign(t)
        return t, b
    
class HashMapping_HyP2(nn.Module):
    # Proxy-Guided Decompositional Hashing (PGDH)
    def __init__(self, input_dim: int, hidden_dim: int, output_dim: int, dropout: float, num_classes: int):
        super().__init__()
        self.proxies = torch.nn.Parameter(torch.randn(num_classes, output_dim))
        nn.init.kaiming_normal_(self.proxies, mode = 'fan_out')
        self.hash_layers = nn.Sequential(
            nn.Linear(input_dim, hidden_dim),
            nn.ReLU(inplace=True),
            # nn.Dropout(p=dropout),
            nn.Linear(hidden_dim, hidden_dim),
            nn.ReLU(inplace=True),
            # nn.Dropout(p=dropout),
            nn.Linear(hidden_dim, output_dim),
        )
        # self.scale = nn.Parameter(torch.Tensor([1.0]))
        # self.last_layer = nn.Tanh()
        self.weights = self.init_weights()
        wb = openpyxl.load_workbook(os.path.join(os.path.dirname(os.path.dirname(__file__)), 'codetable.xlsx'))
        sheet = wb.active
        self.threshold = sheet.cell(row=output_dim+1, column=math.ceil(math.log(num_classes, 2))+1).value

    def init_weights(self):
        for m in self.modules():
            if isinstance(m, nn.Linear):
                nn.init.xavier_normal_(m.weight)
                nn.init.constant_(m.bias, 0.0)

    def forward(self, text_features):
        t = self.hash_layers(text_features)
        b = torch.sign(t)
        return t, b
    
    def cal_loss(self, t, labels, beta = 0.5):
        P_one_hot = labels

        cos = F.normalize(t, p = 2, dim = 1).mm(F.normalize(self.proxies, p = 2, dim = 1).T)
        pos = 1 - cos
        neg = F.relu(cos - self.threshold)

        P_num = len(P_one_hot.nonzero())
        N_num = len((P_one_hot == 0).nonzero())
        L_proxy = torch.where(P_one_hot  ==  1, pos.to(torch.float32), torch.zeros_like(cos).to(torch.float32)).sum() / P_num
        L_neg = torch.where(P_one_hot  ==  0, neg.to(torch.float32), torch.zeros_like(cos).to(torch.float32)).sum() / N_num
        if beta > 0:
            index = labels.sum(dim = 1) > 1
            y_ = labels[index].float()
            x_ = t[index]
            cos_sim = y_.mm(y_.T)
            if len((cos_sim == 0).nonzero()) == 0:
                reg_term = 0
            else:
                x_sim = F.normalize(x_, p = 2, dim = 1).mm(F.normalize(x_, p = 2, dim = 1).T)
                neg = beta * F.relu(x_sim - self.threshold)
                reg_term = torch.where(cos_sim == 0, neg, torch.zeros_like(x_sim)).sum() / len((cos_sim == 0).nonzero())
        else:
            reg_term = 0

        return L_proxy + L_neg + reg_term
    
class InformationLayer(nn.Module):
    def __init__(self, input_dim: int):
        super().__init__()
        self.fc_1 = nn.Sequential(
            nn.Linear(input_dim, input_dim),
            nn.ReLU()
        )
        self.fc_2 = nn.Sequential(
            nn.Linear(input_dim, input_dim),
            nn.ReLU()
        )
        self.weights = self.init_weights()

    def init_weights(self):
        for m in self.modules():
            if isinstance(m, nn.Linear):
                nn.init.xavier_normal_(m.weight)
                nn.init.constant_(m.bias, 0.0)
    
    def forward(self, features, original_features):
        features = self.fc_1(features)
        features = features + original_features
        features = self.fc_2(features)
        return features
    
class VerificationLayer(nn.Module):
    def __init__(self, input_dim: int):
        super().__init__()
        self.fc = nn.Sequential(
            nn.Linear(input_dim, 16),
            nn.BatchNorm1d(16)
        )
        self.weights = self.init_weights()

    def init_weights(self):
        for m in self.modules():
            if isinstance(m, nn.Linear):
                nn.init.xavier_normal_(m.weight)
                nn.init.constant_(m.bias, 0.0)
    
    def forward(self, features):
        features = self.fc(features)
        return features


class HashMapping_CVH(nn.Module):
    # Proxy-Guided Decompositional Hashing (PGDH) with Serial Hash Mapping
    def __init__(self, input_dim: int, output_dim: int, num_classes: int):
        super().__init__()
        self.proxies = torch.nn.Parameter(torch.randn(num_classes, output_dim))
        nn.init.kaiming_normal_(self.proxies, mode = 'fan_out')
        self.information_layers = nn.ModuleList([InformationLayer(input_dim) for i in range(output_dim//16)])
        self.verification_layers = nn.ModuleList([VerificationLayer(input_dim) for i in range(output_dim//16)])
        self.weights = self.init_weights()
        wb = openpyxl.load_workbook(os.path.join(os.path.dirname(os.path.dirname(__file__)), 'codetable.xlsx'))
        sheet = wb.active
        self.threshold = sheet.cell(row=output_dim+1, column=math.ceil(math.log(num_classes, 2))+1).value

    def init_weights(self):
        for m in self.modules():
            if isinstance(m, nn.Linear):
                nn.init.xavier_normal_(m.weight)
                nn.init.constant_(m.bias, 0.0)

    def forward(self, video_features):
        information_codes = video_features
        for idx, (information, verification) in enumerate(zip(self.information_layers, self.verification_layers)):
            information_codes = information(information_codes, video_features)
            partial_hash_codes = verification(information_codes)
            if idx == 0:
                t = partial_hash_codes
            else:
                t = torch.cat([t, partial_hash_codes], dim=-1)
        b = torch.sign(t)
        return t, b
    
    def cal_loss(self, t, labels, beta = 0.5):
        P_one_hot = labels

        cos = F.normalize(t, p = 2, dim = 1).mm(F.normalize(self.proxies, p = 2, dim = 1).T)
        pos = 1 - cos
        neg = F.relu(cos - self.threshold)

        P_num = len(P_one_hot.nonzero())
        N_num = len((P_one_hot == 0).nonzero())
        L_proxy = torch.where(P_one_hot  ==  1, pos.to(torch.float32), torch.zeros_like(cos).to(torch.float32)).sum() / P_num
        L_neg = torch.where(P_one_hot  ==  0, neg.to(torch.float32), torch.zeros_like(cos).to(torch.float32)).sum() / N_num
        if beta > 0:
            index = labels.sum(dim = 1) > 1
            y_ = labels[index].float()
            x_ = t[index]
            cos_sim = y_.mm(y_.T)
            if len((cos_sim == 0).nonzero()) == 0:
                reg_term = 0
            else:
                x_sim = F.normalize(x_, p = 2, dim = 1).mm(F.normalize(x_, p = 2, dim = 1).T)
                neg = beta * F.relu(x_sim - self.threshold)
                reg_term = torch.where(cos_sim == 0, neg, torch.zeros_like(x_sim)).sum() / len((cos_sim == 0).nonzero())
        else:
            reg_term = 0

        return L_proxy + L_neg + reg_term
    
if __name__ == "__main__":
    input_size = 512
    model = HashMapping(input_size, input_size, 64, 0.5)
    param = model.parameters()
    text_features = torch.randn([5, 512]).to(device='cpu')
    t, b = model(text_features)
    print("Binary Code Size: ", b.size())


